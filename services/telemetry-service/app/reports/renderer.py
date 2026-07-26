from __future__ import annotations

import csv
import io
import json
import zipfile
from collections.abc import Iterable, Mapping
from dataclasses import dataclass
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from app.reports.domain import ArtifactDescriptor, sha256_hex

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XLSX_RENDERER_VERSION = "nexolab-xlsx-v1"
_REQUIRED_ARTIFACTS = (
    "manifest.json",
    "source-snapshot.json",
    "telemetry.csv",
    "alert-transitions.csv",
)
_FIXED_ZIP_TIMESTAMP = (1980, 1, 1, 0, 0, 0)
_EXCEL_CELL_LIMIT = 32_767
_EXCEL_CELL_CHUNK = 32_000


class ReportRenderError(ValueError):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code


@dataclass(frozen=True, slots=True)
class RenderedReportArtifact:
    descriptor: ArtifactDescriptor
    content: bytes
    renderer_version: str


@dataclass(frozen=True, slots=True)
class ParsedCsv:
    columns: tuple[str, ...]
    rows: list[dict[str, str]]


def render_xlsx_report(artifacts: Mapping[str, bytes]) -> RenderedReportArtifact:
    frozen = {name: bytes(content) for name, content in artifacts.items()}
    _require_artifacts(frozen)

    manifest = _json_object(frozen["manifest.json"], "manifest.json")
    source = _json_object(frozen["source-snapshot.json"], "source-snapshot.json")
    descriptors = _verify_evidence(frozen, manifest, source)
    telemetry = _csv_rows(
        frozen["telemetry.csv"],
        "telemetry.csv",
        descriptors["telemetry.csv"].row_count,
    )
    alerts = _csv_rows(
        frozen["alert-transitions.csv"],
        "alert-transitions.csv",
        descriptors["alert-transitions.csv"].row_count,
    )

    workbook = _build_workbook(
        manifest=manifest,
        source=source,
        descriptors=descriptors,
        manifest_content=frozen["manifest.json"],
        telemetry=telemetry,
        alerts=alerts,
    )
    raw = io.BytesIO()
    workbook.save(raw)
    content = _normalize_xlsx_archive(raw.getvalue())
    descriptor = ArtifactDescriptor.from_bytes(
        name="report.xlsx",
        media_type=XLSX_MEDIA_TYPE,
        content=content,
    )
    return RenderedReportArtifact(descriptor, content, XLSX_RENDERER_VERSION)


def _require_artifacts(artifacts: Mapping[str, bytes]) -> None:
    missing = [name for name in _REQUIRED_ARTIFACTS if name not in artifacts]
    if missing:
        raise ReportRenderError(
            "report_artifact_missing",
            f"required immutable report artifacts are missing: {', '.join(missing)}",
        )


def _json_object(content: bytes, name: str) -> dict[str, Any]:
    try:
        value = json.loads(content.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as error:
        raise ReportRenderError(
            "report_artifact_invalid_json",
            f"{name} is not valid UTF-8 JSON",
        ) from error
    if not isinstance(value, dict):
        raise ReportRenderError(
            "report_artifact_invalid_json",
            f"{name} must contain a JSON object",
        )
    return value


def _verify_evidence(
    artifacts: Mapping[str, bytes],
    manifest: Mapping[str, Any],
    source: Mapping[str, Any],
) -> dict[str, ArtifactDescriptor]:
    report = _mapping(manifest.get("report"), "manifest.report")
    source_digest = _required_string(report.get("source_sha256"), "source_sha256")
    if sha256_hex(artifacts["source-snapshot.json"]) != source_digest:
        raise ReportRenderError(
            "report_source_digest_mismatch",
            "source-snapshot.json does not match the immutable manifest source digest",
        )

    raw_descriptors = manifest.get("artifacts")
    if not isinstance(raw_descriptors, list):
        raise ReportRenderError(
            "report_manifest_invalid",
            "manifest.artifacts must be an array",
        )

    descriptors: dict[str, ArtifactDescriptor] = {}
    for index, value in enumerate(raw_descriptors):
        item = _mapping(value, f"manifest.artifacts[{index}]")
        try:
            descriptor = ArtifactDescriptor(
                name=_required_string(item.get("name"), "artifact.name"),
                media_type=_required_string(
                    item.get("media_type"),
                    "artifact.media_type",
                ),
                sha256=_required_string(item.get("sha256"), "artifact.sha256"),
                size_bytes=_required_int(
                    item.get("size_bytes"),
                    "artifact.size_bytes",
                ),
                row_count=_optional_int(
                    item.get("row_count"),
                    "artifact.row_count",
                ),
            )
        except ValueError as error:
            raise ReportRenderError(
                "report_manifest_invalid",
                f"invalid descriptor at manifest.artifacts[{index}]",
            ) from error
        if descriptor.name in descriptors:
            raise ReportRenderError(
                "report_manifest_invalid",
                f"duplicate artifact descriptor: {descriptor.name}",
            )
        descriptors[descriptor.name] = descriptor

    for name in ("source-snapshot.json", "telemetry.csv", "alert-transitions.csv"):
        if name not in descriptors:
            raise ReportRenderError(
                "report_manifest_invalid",
                f"manifest has no descriptor for {name}",
            )

    for name, descriptor in descriptors.items():
        content = artifacts.get(name)
        if content is None:
            raise ReportRenderError(
                "report_artifact_missing",
                f"manifest artifact is missing: {name}",
            )
        if (
            descriptor.sha256 != sha256_hex(content)
            or descriptor.size_bytes != len(content)
        ):
            raise ReportRenderError(
                "report_artifact_digest_mismatch",
                f"{name} does not match its immutable manifest descriptor",
            )

    if (
        report.get("organization_id") != source.get("organization_id")
        or report.get("session_id") != source.get("session_id")
    ):
        raise ReportRenderError(
            "report_source_identity_mismatch",
            "manifest report identity does not match the frozen source snapshot",
        )

    source_evidence = _mapping(source.get("evidence"), "source.evidence")
    source_names = {
        "telemetry": "telemetry.csv",
        "alert_transitions": "alert-transitions.csv",
    }
    for source_key, artifact_name in source_names.items():
        evidence = _mapping(source_evidence.get(source_key), f"source.evidence.{source_key}")
        descriptor = descriptors[artifact_name]
        if (
            evidence.get("name") != descriptor.name
            or evidence.get("sha256") != descriptor.sha256
            or evidence.get("size_bytes") != descriptor.size_bytes
            or evidence.get("row_count") != descriptor.row_count
        ):
            raise ReportRenderError(
                "report_source_descriptor_mismatch",
                f"source evidence descriptor does not match {artifact_name}",
            )
    return descriptors


def _csv_rows(content: bytes, name: str, expected_count: int | None) -> ParsedCsv:
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError as error:
        raise ReportRenderError(
            "report_artifact_invalid_csv",
            f"{name} is not valid UTF-8 CSV",
        ) from error
    try:
        reader = csv.DictReader(io.StringIO(text, newline=""))
        if not reader.fieldnames or any(not field for field in reader.fieldnames):
            raise ReportRenderError(
                "report_artifact_invalid_csv",
                f"{name} has no valid header",
            )
        if len(reader.fieldnames) != len(set(reader.fieldnames)):
            raise ReportRenderError(
                "report_artifact_invalid_csv",
                f"{name} contains duplicate columns",
            )
        rows = [dict(row) for row in reader]
    except csv.Error as error:
        raise ReportRenderError(
            "report_artifact_invalid_csv",
            f"{name} cannot be parsed as CSV",
        ) from error
    if expected_count is not None and len(rows) != expected_count:
        raise ReportRenderError(
            "report_artifact_row_count_mismatch",
            f"{name} row count does not match its immutable descriptor",
        )
    return ParsedCsv(tuple(reader.fieldnames), rows)


def _build_workbook(
    *,
    manifest: Mapping[str, Any],
    source: Mapping[str, Any],
    descriptors: Mapping[str, ArtifactDescriptor],
    manifest_content: bytes,
    telemetry: ParsedCsv,
    alerts: ParsedCsv,
) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "NEXOLAB"
    workbook.properties.lastModifiedBy = "NEXOLAB"
    workbook.properties.title = "NEXOLAB immutable test report"
    workbook.properties.subject = "Frozen telemetry and alert evidence"
    workbook.properties.description = XLSX_RENDERER_VERSION
    workbook.properties.keywords = "NEXOLAB,report,telemetry,evidence"
    workbook.properties.category = "Test report"
    fixed_time = datetime(1980, 1, 1)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time
    workbook.calculation.calcMode = "manual"
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False

    summary_sheet = workbook.create_sheet("Summary")
    telemetry_sheet = workbook.create_sheet("Telemetry")
    alerts_sheet = workbook.create_sheet("Alerts")
    evidence_sheet = workbook.create_sheet("Evidence")

    _populate_summary(
        summary_sheet,
        manifest,
        source,
        telemetry.rows,
        alerts.rows,
    )
    _populate_table(
        telemetry_sheet,
        telemetry.columns,
        telemetry.rows,
        numeric_columns={"value"},
    )
    _populate_table(alerts_sheet, alerts.columns, alerts.rows)
    _populate_evidence(evidence_sheet, descriptors, manifest_content)
    workbook.active = 0
    return workbook


def _populate_summary(
    sheet: Worksheet,
    manifest: Mapping[str, Any],
    source: Mapping[str, Any],
    telemetry_rows: list[dict[str, str]],
    alert_rows: list[dict[str, str]],
) -> None:
    report = _mapping(manifest.get("report"), "manifest.report")
    metadata = _mapping(source.get("metadata"), "source.metadata")
    session = _mapping(metadata.get("session"), "source.metadata.session")
    rows: list[tuple[str, Any]] = [
        ("Renderer version", XLSX_RENDERER_VERSION),
        ("Report ID", report.get("id")),
        ("Organization ID", report.get("organization_id")),
        ("Session ID", report.get("session_id")),
        ("Report version", report.get("version")),
        ("Generated at (UTC)", report.get("generated_at")),
        ("Generated by", report.get("generated_by")),
        ("Generator version", report.get("generator_version")),
        ("Source SHA-256", report.get("source_sha256")),
        ("Source started at (UTC)", source.get("source_started_at")),
        ("Source ended at (UTC)", source.get("source_ended_at")),
        ("Session state", session.get("state")),
        ("Session title", session.get("title")),
        ("Customer", session.get("customer")),
        ("Test object", session.get("test_object")),
        ("Model", session.get("model")),
        ("Serial number", session.get("serial_number")),
        ("Standard", session.get("standard")),
        ("Method", session.get("method")),
        ("Telemetry rows", len(telemetry_rows)),
        ("Alert transition rows", len(alert_rows)),
    ]
    sheet.append(("Field", "Value"))
    for key, value in rows:
        sheet.append((key, _cell_value(value)))
    sheet.append(())
    sheet.append(("Frozen metadata path", "Value"))
    for path, value in _flatten_metadata(metadata):
        text = _metadata_text(value)
        for index, chunk in enumerate(_chunks(text), start=1):
            suffix = f"#part-{index}" if len(text) > _EXCEL_CELL_LIMIT else ""
            sheet.append((f"metadata.{path}{suffix}", chunk))
    _format_header(sheet, 1)
    _format_header(sheet, len(rows) + 3)
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 96


def _populate_table(
    sheet: Worksheet,
    columns: tuple[str, ...],
    rows: list[dict[str, str]],
    *,
    numeric_columns: set[str] | None = None,
) -> None:
    numeric = numeric_columns or set()
    headers = list(columns)
    if headers:
        sheet.append(headers)
        for row in rows:
            sheet.append(
                [
                    _table_value(row.get(header, ""), header in numeric)
                    for header in headers
                ]
            )
        _format_header(sheet, 1)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, header in enumerate(headers, start=1):
            sheet.column_dimensions[sheet.cell(1, index).column_letter].width = max(14, min(32, len(header) + 3))
    else:
        sheet.append(("No evidence rows",))
        _format_header(sheet, 1)


def _populate_evidence(
    sheet: Worksheet,
    descriptors: Mapping[str, ArtifactDescriptor],
    manifest_content: bytes,
) -> None:
    sheet.append(("Artifact", "Media type", "SHA-256", "Size bytes", "Row count", "Verified"))
    for descriptor in sorted(descriptors.values(), key=lambda item: item.name):
        sheet.append(
            (
                descriptor.name,
                descriptor.media_type,
                descriptor.sha256,
                descriptor.size_bytes,
                descriptor.row_count,
                "yes",
            )
        )
    sheet.append(
        (
            "manifest.json",
            "application/json",
            sha256_hex(manifest_content),
            len(manifest_content),
            None,
            "yes",
        )
    )
    _format_header(sheet, 1)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = (28, 44, 68, 16, 14, 12)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width


def _format_header(sheet: Worksheet, row: int) -> None:
    for cell in sheet[row]:
        cell.font = Font(bold=True)


def _flatten_metadata(value: Any, prefix: str = "") -> Iterable[tuple[str, Any]]:
    if isinstance(value, Mapping):
        for key in sorted(value):
            child = f"{prefix}.{key}" if prefix else str(key)
            yield from _flatten_metadata(value[key], child)
        return
    if isinstance(value, list):
        for index, item in enumerate(value):
            child = f"{prefix}[{index}]"
            yield from _flatten_metadata(item, child)
        if not value:
            yield prefix, "[]"
        return
    yield prefix, value


def _chunks(value: str) -> list[str]:
    if len(value) <= _EXCEL_CELL_LIMIT:
        return [value]
    return [value[index : index + _EXCEL_CELL_CHUNK] for index in range(0, len(value), _EXCEL_CELL_CHUNK)]


def _cell_value(value: Any) -> str | int | float:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float, str)):
        return value
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _metadata_text(value: Any) -> str:
    converted = _cell_value(value)
    return str(converted)


def _table_value(value: str, numeric: bool) -> str | float:
    if numeric and value:
        try:
            return float(value)
        except ValueError as error:
            raise ReportRenderError(
                "report_artifact_invalid_csv",
                f"numeric evidence value is invalid: {value}",
            ) from error
    return value


def _normalize_xlsx_archive(content: bytes) -> bytes:
    source = io.BytesIO(content)
    target = io.BytesIO()
    try:
        with zipfile.ZipFile(source, "r") as archive:
            names = sorted(archive.namelist())
            if len(names) != len(set(names)):
                raise ReportRenderError(
                    "xlsx_archive_invalid",
                    "generated XLSX contains duplicate ZIP entries",
                )
            with zipfile.ZipFile(
                target,
                "w",
                compression=zipfile.ZIP_DEFLATED,
                compresslevel=9,
                strict_timestamps=True,
            ) as normalized:
                for name in names:
                    info = zipfile.ZipInfo(name, date_time=_FIXED_ZIP_TIMESTAMP)
                    info.compress_type = zipfile.ZIP_DEFLATED
                    info.create_system = 3
                    info.external_attr = 0o600 << 16
                    normalized.writestr(info, archive.read(name), compress_type=zipfile.ZIP_DEFLATED, compresslevel=9)
    except zipfile.BadZipFile as error:
        raise ReportRenderError(
            "xlsx_archive_invalid",
            "generated XLSX is not a valid ZIP archive",
        ) from error
    return target.getvalue()


def _mapping(value: Any, field: str) -> Mapping[str, Any]:
    if not isinstance(value, Mapping):
        raise ReportRenderError("report_manifest_invalid", f"{field} must be an object")
    return value


def _required_string(value: Any, field: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ReportRenderError("report_manifest_invalid", f"{field} must be a non-empty string")
    return value


def _required_int(value: Any, field: str) -> int:
    if not isinstance(value, int) or isinstance(value, bool) or value < 0:
        raise ReportRenderError("report_manifest_invalid", f"{field} must be a nonnegative integer")
    return value


def _optional_int(value: Any, field: str) -> int | None:
    if value is None:
        return None
    return _required_int(value, field)

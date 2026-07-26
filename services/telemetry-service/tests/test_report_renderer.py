from __future__ import annotations

import hashlib
import io
import json
import zipfile

import pytest
from openpyxl import load_workbook

from app.reports.renderer import (
    XLSX_MEDIA_TYPE,
    XLSX_RENDERER_VERSION,
    ReportRenderError,
    render_xlsx_report,
)


def _canonical(value: object) -> bytes:
    return (
        json.dumps(
            value,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
        + b"\n"
    )


def _descriptor(
    name: str,
    media_type: str,
    content: bytes,
    row_count: int | None,
) -> dict[str, object]:
    return {
        "name": name,
        "media_type": media_type,
        "sha256": hashlib.sha256(content).hexdigest(),
        "size_bytes": len(content),
        "row_count": row_count,
    }


def report_artifacts() -> dict[str, bytes]:
    telemetry = (
        "event_id,captured_at,node_id,equipment_id,channel_id,metric,value,unit,"
        "quality,alarm,source,session_id,stage_id,binding_id,config_snapshot_id\n"
        "event-1,2026-07-26T12:00:01.000000Z,edge-01,K106,106-03,"
        "temperature.probe,8.7,degC,valid,,reports-test,session-1,stage-1,"
        "binding-1,snapshot-1\n"
        "event-2,2026-07-26T12:00:02.000000Z,edge-01,K106,106-03,"
        "temperature.probe,9.4,degC,valid,,reports-test,session-1,stage-1,"
        "binding-1,snapshot-1\n"
    ).encode("utf-8")
    alerts = (
        "alert_id,transition_id,rule_id,rule_version_id,event_type,"
        "previous_state,next_state,actor_id,actor_source,reason,occurred_at,"
        "severity,node_id,equipment_id,channel_id,metric,session_id,stage_id,"
        "binding_id\n"
        "alert-1,transition-1,rule-1,rule-version-1,alert_acknowledged,active,"
        "acknowledged,manager-1,verified-jwt,Operator inspected equipment,"
        "2026-07-26T12:01:00.000000Z,critical,edge-01,K106,106-03,"
        "temperature.probe,session-1,stage-1,binding-1\n"
    ).encode("utf-8")
    telemetry_descriptor = _descriptor(
        "telemetry.csv",
        "text/csv; charset=utf-8",
        telemetry,
        2,
    )
    alerts_descriptor = _descriptor(
        "alert-transitions.csv",
        "text/csv; charset=utf-8",
        alerts,
        1,
    )
    source = _canonical(
        {
            "schema": "nexolab.report-source.v1",
            "organization_id": "organization-1",
            "session_id": "session-1",
            "source_started_at": "2026-07-26T12:00:00.000000Z",
            "source_ended_at": "2026-07-26T13:00:00.000000Z",
            "metadata": {
                "session": {
                    "state": "completed",
                    "title": "Cold display compliance test",
                    "customer": "NEXOLAB",
                    "test_object": "K106",
                    "model": "Showcase 106",
                    "serial_number": "SN-106",
                    "standard": "EN 23953",
                    "method": "stabilization",
                    "revision": 3,
                },
                "notes": [],
            },
            "evidence": {
                "telemetry": telemetry_descriptor,
                "alert_transitions": alerts_descriptor,
            },
        }
    )
    source_descriptor = _descriptor(
        "source-snapshot.json",
        "application/json",
        source,
        None,
    )
    manifest = _canonical(
        {
            "schema": "nexolab.report-manifest.v1",
            "report": {
                "id": "report-1",
                "organization_id": "organization-1",
                "session_id": "session-1",
                "version": 1,
                "source_sha256": hashlib.sha256(source).hexdigest(),
                "generated_at": "2026-07-26T14:00:00.000000Z",
                "generated_by": "engineer-1",
                "generator_version": "reports-domain-v1",
            },
            "artifacts": [
                alerts_descriptor,
                source_descriptor,
                telemetry_descriptor,
            ],
        }
    )
    return {
        "manifest.json": manifest,
        "source-snapshot.json": source,
        "telemetry.csv": telemetry,
        "alert-transitions.csv": alerts,
    }


def test_xlsx_render_is_byte_stable_and_archive_normalized() -> None:
    artifacts = report_artifacts()

    first = render_xlsx_report(artifacts)
    second = render_xlsx_report(dict(reversed(list(artifacts.items()))))

    assert first.content == second.content
    assert first.descriptor.media_type == XLSX_MEDIA_TYPE
    assert first.descriptor.sha256 == hashlib.sha256(first.content).hexdigest()
    assert first.renderer_version == XLSX_RENDERER_VERSION
    with zipfile.ZipFile(io.BytesIO(first.content)) as archive:
        assert archive.namelist() == sorted(archive.namelist())
        assert {entry.date_time for entry in archive.infolist()} == {
            (1980, 1, 1, 0, 0, 0)
        }


def test_xlsx_contains_frozen_summary_telemetry_alerts_and_evidence() -> None:
    rendered = render_xlsx_report(report_artifacts())

    workbook = load_workbook(io.BytesIO(rendered.content), data_only=True)

    assert workbook.sheetnames == ["Summary", "Telemetry", "Alerts", "Evidence"]
    assert workbook["Summary"]["B2"].value == XLSX_RENDERER_VERSION
    assert workbook["Summary"]["B3"].value == "report-1"
    assert workbook["Telemetry"]["A2"].value == "event-1"
    assert workbook["Telemetry"]["G2"].value == 8.7
    assert workbook["Alerts"]["I2"].value == "verified-jwt"
    assert workbook["Evidence"]["F2"].value == "yes"
    summary_values = [cell.value for cell in workbook["Summary"]["B"]]
    assert "3" in summary_values


def test_xlsx_render_rejects_tampered_immutable_evidence() -> None:
    artifacts = report_artifacts()
    artifacts["telemetry.csv"] += b"tampered"

    with pytest.raises(ReportRenderError) as captured:
        render_xlsx_report(artifacts)

    assert captured.value.code == "report_artifact_digest_mismatch"


def test_xlsx_render_rejects_missing_immutable_artifact() -> None:
    artifacts = report_artifacts()
    artifacts.pop("alert-transitions.csv")

    with pytest.raises(ReportRenderError) as captured:
        render_xlsx_report(artifacts)

    assert captured.value.code == "report_artifact_missing"
